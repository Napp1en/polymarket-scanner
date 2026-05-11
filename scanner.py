import requests
import pandas as pd
from datetime import datetime
import json

# Einstellungen
BANKROLL = 100
TOP_N = 5
MIN_RAW_ROI = 0.05
OUTFILE = "polymarket_cs_scanner.xlsx"

EVENTS_URL = "https://gamma-api.polymarket.com/events"

KEYWORDS = [
    "counter-strike",
    "counter strike",
    "cs2",
    "pgl",
    "pgl astana",
    "pgl bucharest",
    "pgl major",
    "blast premier",
    "iem cologne",
    "iem katowice",
    "esl pro league",
    "fissure playground",
]

TEAMS = [
    "falcons", "spirit", "furia", "g2", "navi", "vitality",
    "mouz", "faze", "astralis", "liquid", "virtus.pro",
    "the mongolz", "pain", "heroic"
]


def parse_list(value):
    if isinstance(value, list):
        return value

    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return []

    return []

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def format_excel(outfile):
    from openpyxl import load_workbook

    wb = load_workbook(outfile)
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 🎨 Header
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill

    # 📏 Spaltenbreite
    widths = {
        "A": 20,
        "B": 35,  # Event
        "C": 50,
        "D": 12,
        "E": 14,
        "F": 10,
        "G": 10,
        "H": 11,
        "I": 10,
        "J": 14,
        "K": 50,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 🎨 Farben für Events
    event_colors = [
        "FFF2CC",  # hellgelb
        "DDEBF7",  # hellblau
        "E2EFDA",  # hellgrün
        "FCE4D6",  # orange
        "E4DFEC",  # lila
        "F8CBAD",  # rot/rosa
    ]

    event_column = 2  # Spalte B = Event
    event_map = {}
    color_index = 0

    for row in range(2, ws.max_row + 1):
        event_name = ws.cell(row=row, column=event_column).value

        if event_name not in event_map:
            event_map[event_name] = event_colors[color_index % len(event_colors)]
            color_index += 1

        fill = PatternFill("solid", fgColor=event_map[event_name])

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    # 💰 Zahlenformate
    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].number_format = "0.0000"
        ws[f"E{row}"].number_format = "0.0000"
        ws[f"F{row}"].number_format = "0.00"
        ws[f"G{row}"].number_format = "$0.00"
        ws[f"H{row}"].number_format = "$0.00"
        ws[f"I{row}"].number_format = "$0.00"

    wb.save(outfile)

def get_events(limit=250, max_pages=4):
    events_by_id = {}

    orders = [
        "volume_24hr",
        "volume",
        "liquidity",
    ]

    for order in orders:
        print(f"Lade Events nach: {order}")

        for page in range(max_pages):
            params = {
                "active": "true",
                "closed": "false",
                "limit": limit,
                "offset": page * limit,
                "order": order,
                "ascending": "false",
            }

            r = requests.get(EVENTS_URL, params=params, timeout=10)
            r.raise_for_status()

            batch = r.json()

            if not batch:
                break

            for e in batch:
                event_id = e.get("id") or e.get("slug")
                events_by_id[event_id] = e

            print(f"  Seite {page + 1}: {len(batch)} Events")

            if len(batch) < limit:
                break

    return list(events_by_id.values())


def is_cs_event(event):
    text = " ".join([
        str(event.get("title", "")),
        str(event.get("slug", "")),
        str(event.get("description", "")),
        str(event.get("seriesSlug", "")),
        str(event.get("category", "")),
    ]).lower()

    return any(k in text for k in KEYWORDS)


def is_team_market(market):
    text = " ".join([
        str(market.get("question", "")),
        str(market.get("title", "")),
        str(market.get("description", "")),
        str(market.get("slug", "")),
    ]).lower()

    return any(team in text for team in TEAMS)


def scan():
    rows = []
    events = get_events()

    print(f"Events geladen: {len(events)}")

    cs_events = [e for e in events if is_cs_event(e)]
    print(f"CS Events gefunden: {len(cs_events)}")

    for event in cs_events:
        markets = event.get("markets", [])

        team_prices = []

        for market in markets:
            outcomes = parse_list(market.get("outcomes"))
            prices = parse_list(market.get("outcomePrices"))

            if not outcomes or not prices:
                continue

            # Polymarket Team-Märkte sind meistens: ["Yes", "No"]
            # Wir brauchen nur den YES-Preis
            yes_price = None

            for outcome, price in zip(outcomes, prices):
                if str(outcome).lower() == "yes":
                    yes_price = float(price)

            if yes_price is None or yes_price <= 0:
                continue

            team_name = (
                market.get("question")
                or market.get("title")
                or market.get("slug")
            )

            team_prices.append({
                "team": team_name,
                "price": yes_price,
                "market": market,
            })

        if len(team_prices) < TOP_N:
            continue

        # Top 5 nach höchstem YES-Preis
        team_prices.sort(key=lambda x: x["price"], reverse=True)
        top_teams = team_prices[:TOP_N]

        yes_sum = sum(t["price"] for t in top_teams)

        if yes_sum <= 0:
            continue

        roi = (1 / yes_sum) - 1

        # Nur positive / gewünschte ROI-Spots
        if roi < MIN_RAW_ROI:
            continue

        payout = BANKROLL / yes_sum
        profit = payout - BANKROLL

        for t in top_teams:
            price = t["price"]
            stake = (price / yes_sum) * BANKROLL

            rows.append({
                "Zeit": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Event": event.get("title"),
                "Team / Markt": t["team"],
                "YES Preis": round(price, 4),
                f"Summe Top {TOP_N}": round(yes_sum, 4),
                "ROI %": round(roi * 100, 2),
                "Stake $": round(stake, 2),
                "Payout $": round(payout, 2),
                "Profit $": round(profit, 2),
                "Volumen Event": event.get("volume"),
                "Link": "https://polymarket.com/event/" + str(event.get("slug", "")),
            })

    if not rows:
        print("Keine passenden Spots gefunden.")
        return

    df = pd.DataFrame(rows)
    df = df.sort_values(by="ROI %", ascending=False)
    df.to_excel(OUTFILE, index=False)
    format_excel(OUTFILE)

    print(f"✅ Fertig: {OUTFILE}")
    print(df.head(20))


if __name__ == "__main__":
    scan()
        